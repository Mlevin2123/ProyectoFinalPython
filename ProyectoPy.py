import os
import tkinter as tk
from tkinter import ttk, messagebox, Menu
import csv
from fpdf import FPDF
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import ttkbootstrap as ttkb  # Mejora visual
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

class LibretaContactosApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Libreta de Contactos V2.0')
        self.root.geometry("+350+80")
        self.root.resizable(0, 0)

        # Menú principal
        menu = Menu(self.root)
        self.root.config(menu=menu)
        archivo_menu = Menu(menu, tearoff=0, bg="#FFBB20")
        archivo_menu.add_command(label="Mostrar todos los contactos", command=self.mostrar_contactos)
        archivo_menu.add_command(label="Cerrar", command=self.root.quit)
        menu.add_cascade(label="Menú", menu=archivo_menu)

        # Marco de entrada de datos
        entrada_frame = tk.LabelFrame(self.root, bg="#53CDB8")
        entrada_frame.grid(row=0, column=0)

        tk.Label(entrada_frame, text='Nombre', bg="#53CDB8").grid(row=0, column=0)
        self.nombre_entry = tk.Entry(entrada_frame, width=25)
        self.nombre_entry.grid(row=1, column=0)

        tk.Label(entrada_frame, text='Teléfono', bg="#53CDB8").grid(row=0, column=1)
        self.telefono_entry = tk.Entry(entrada_frame, width=20)
        self.telefono_entry.grid(row=1, column=1)

        tk.Label(entrada_frame, text='Correo', bg="#53CDB8").grid(row=0, column=2)
        self.correo_entry = tk.Entry(entrada_frame, width=25)
        self.correo_entry.grid(row=1, column=2)

        # Botones de acciones
        acciones_frame = ttkb.Frame(self.root)
        acciones_frame.grid(row=1, column=0, pady=10)

        ttkb.Button(acciones_frame, text="Agregar Contacto", command=self.agregar_contacto, bootstyle="success").grid(row=0, column=0, padx=5)
        ttkb.Button(acciones_frame, text="Buscar Contacto", command=self.buscar_contacto, bootstyle="info").grid(row=0, column=1, padx=5)
        ttkb.Button(acciones_frame, text="Eliminar Contacto", command=self.eliminar_contacto, bootstyle="danger").grid(row=0, column=2, padx=5)
        ttkb.Button(acciones_frame, text="Editar", command=self.editar_contacto, bootstyle="warning").grid(row=0, column=3, padx=5)
        ttkb.Button(acciones_frame, text="Descargar Excel", command=self.descargar_csv, bootstyle="secondary").grid(row=0, column=4, padx=5)
        ttkb.Button(acciones_frame, text="Descargar PDF", command=self.descargar_pdf, bootstyle="secondary").grid(row=0, column=5, padx=5)
        ttkb.Button(acciones_frame, text="Enviar Correo", command=self.enviar_correo_contacto, bootstyle="primary").grid(row=0, column=6, padx=5)

        # Tabla de contactos
        self.tabla = ttkb.Treeview(self.root, columns=("Teléfono", "Correo"), height=15)
        self.tabla.grid(row=2, column=0, padx=10, pady=10)
        self.tabla.heading("#0", text="Nombre")
        self.tabla.heading("Teléfono", text="Teléfono")
        self.tabla.heading("Correo", text="Correo")

        # Scroll para la tabla
        scrollbar_y = ttkb.Scrollbar(self.root, orient="vertical", command=self.tabla.yview)
        scrollbar_y.grid(row=2, column=1, sticky="ns")
        self.tabla.configure(yscroll=scrollbar_y.set)

        # Evento para cargar contacto al seleccionar
        self.tabla.bind("<<TreeviewSelect>>", self.cargar_contacto_seleccionado)



    def agregar_contacto(self):
        nombre, telefono, correo = self.nombre_entry.get(), self.telefono_entry.get(), self.correo_entry.get()
        if not nombre or not telefono or not correo:
            messagebox.showwarning("Datos incompletos", "Por favor, complete todos los campos")
            return
        with open("libreta_contactos.csv", "a", newline='') as f:
            writer = csv.writer(f)
            writer.writerow((nombre, telefono, correo))
        self.tabla.insert("", "end", text=nombre, values=(telefono, correo))
        self.limpiar_entradas()

    def buscar_contacto(self):
        criterio = self.nombre_entry.get().lower()  # Convertir a minúsculas
        if not criterio:
           messagebox.showinfo("Búsqueda vacía", "Por favor, ingrese un criterio para buscar.")
           return

    # Limpiar cualquier selección anterior en la tabla
        self.tabla.selection_remove(self.tabla.selection())

    # Buscar coincidencias
        coincidencias = []
        for item in self.tabla.get_children():
         nombre = self.tabla.item(item, "text").lower()  # Convertir a minúsculas para comparar
         if criterio in nombre:
            coincidencias.append(item)

        if coincidencias:
        # Seleccionar las coincidencias en la tabla
         for item in coincidencias:
            self.tabla.selection_add(item)
        else:
         messagebox.showinfo("Sin resultados", f"No se encontraron contactos que coincidan con '{criterio}'")



    def cargar_contacto_seleccionado(self, event):
        # Obtener el contacto seleccionado en la tabla
        item_seleccionado = self.tabla.selection()
        if not item_seleccionado:
            return

        # Obtener los datos del contacto
        contacto = self.tabla.item(item_seleccionado[0], "text")
        telefono, correo = self.tabla.item(item_seleccionado[0], "values")

        # Cargar los datos en las entradas
        self.nombre_entry.delete(0, tk.END)
        self.nombre_entry.insert(0, contacto)
        self.telefono_entry.delete(0, tk.END)
        self.telefono_entry.insert(0, telefono)
        self.correo_entry.delete(0, tk.END)
        self.correo_entry.insert(0, correo)



    def eliminar_contacto(self):
        nombre = self.nombre_entry.get()
        if not nombre:
            messagebox.showwarning("Campo vacío", "Por favor, ingrese el nombre del contacto a eliminar.")
            return
            
        confirmar = messagebox.askyesno("Eliminar Contacto", f"¿Eliminar el contacto '{nombre}'?")
        if confirmar:
            try:
                contactos_actuales = []
                with open("libreta_contactos.csv", "r", newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if row and row[0].lower() != nombre.lower():
                            contactos_actuales.append(row)

                with open("libreta_contactos.csv", "w", newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerows(contactos_actuales)

                self.mostrar_contactos()
                messagebox.showinfo("Contacto eliminado", f"El contacto '{nombre}' ha sido eliminado.")
            except Exception as e:
                messagebox.showerror("Error", f"Error al eliminar el contacto: {e}")

    def editar_contacto(self):
        # Obtener la selección de la tabla
        item_seleccionado = self.tabla.selection()
        if not item_seleccionado:
            messagebox.showwarning("Selección requerida", "Por favor, seleccione un contacto en la tabla para editar.")
            return

        # Obtener datos actuales del contacto seleccionado
        contacto_actual = self.tabla.item(item_seleccionado, "text")
        telefono_actual, correo_actual = self.tabla.item(item_seleccionado, "values")

        # Cargar los datos en las entradas
        self.nombre_entry.delete(0, tk.END)
        self.nombre_entry.insert(0, contacto_actual)
        self.telefono_entry.delete(0, tk.END)
        self.telefono_entry.insert(0, telefono_actual)
        self.correo_entry.delete(0, tk.END)
        self.correo_entry.insert(0, correo_actual)

        # Confirmar y guardar cambios
        def guardar_cambios():
            nuevo_nombre = self.nombre_entry.get()
            nuevo_telefono = self.telefono_entry.get()
            nuevo_correo = self.correo_entry.get()

            if not nuevo_nombre or not nuevo_telefono or not nuevo_correo:
                messagebox.showwarning("Datos incompletos", "Por favor, complete todos los campos.")
                return

            # Actualizar el archivo CSV
            contactos_actualizados = []
            with open("libreta_contactos.csv", "r", newline='', encoding="utf-8") as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and row[0] == contacto_actual:  # Buscar el contacto original
                        contactos_actualizados.append([nuevo_nombre, nuevo_telefono, nuevo_correo])
                    else:
                        contactos_actualizados.append(row)

            with open("libreta_contactos.csv", "w", newline='', encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(contactos_actualizados)

            # Actualizar la tabla
            self.mostrar_contactos()
            self.limpiar_entradas()
            messagebox.showinfo("Contacto actualizado", f"El contacto '{nuevo_nombre}' ha sido actualizado.")
            editar_ventana.destroy()

        # Crear una ventana emergente para confirmar la edición
        editar_ventana = tk.Toplevel(self.root)
        editar_ventana.title("Confirmar Edición")
        editar_ventana.geometry("300x150")
        editar_ventana.configure(bg="#53CDB8")

        tk.Label(editar_ventana, text="¿Guardar cambios en el contacto?", bg="#53CDB8").pack(pady=20)
        tk.Button(editar_ventana, text="Guardar", command=guardar_cambios, bg="#FFBB20").pack(side=tk.LEFT, padx=20)
        tk.Button(editar_ventana, text="Cancelar", command=editar_ventana.destroy, bg="#F26262").pack(side=tk.RIGHT, padx=20)

    def mostrar_contactos(self):
        for item in self.tabla.get_children():
            self.tabla.delete(item)
        try:
            with open("libreta_contactos.csv", "r", newline='', encoding='utf-8') as f:
                for nombre, telefono, correo in csv.reader(f):
                    self.tabla.insert("", "end", text=nombre, values=(telefono, correo))
        except FileNotFoundError:
            pass

    def limpiar_entradas(self):
        self.nombre_entry.delete(0, tk.END)
        self.telefono_entry.delete(0, tk.END)
        self.correo_entry.delete(0, tk.END)

    def descargar_csv(self):
        try:
            # Definir la ruta de descarga en la carpeta "Descargas"
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            archivo_excel = os.path.join(downloads_path, "libreta_contactos.xlsx")

            # Crear un libro de trabajo y una hoja
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista de los Contactos"

            # Establecer el encabezado
            encabezado = ["Nombre", "Teléfono", "Correo"]
            for col_num, header in enumerate(encabezado, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)  # Títulos en negrita
                cell.alignment = Alignment(horizontal="center")  # Centrar el texto

            # Escribir los datos de los contactos
            with open("libreta_contactos.csv", "r", newline='', encoding="utf-8") as f:
                reader = csv.reader(f)
                for row_num, row in enumerate(reader, start=2):  # Comenzar desde la fila 2
                    for col_num, value in enumerate(row, start=1):
                        ws.cell(row=row_num, column=col_num, value=value)

            # Guardar el archivo Excel en la carpeta de Descargas
            wb.save(archivo_excel)

            # Mostrar mensaje de éxito
            messagebox.showinfo("Excel descargado", f"El archivo Excel ha sido descargado en: {archivo_excel}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al generar el archivo Excel: {e}")

    def descargar_pdf(self):
        try:
            downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
            archivo_pdf = os.path.join(downloads_path, "libreta_contactos.pdf")
        
        # Crear un objeto FPDF
            pdf = FPDF()
            pdf.add_page()
        
        # Establecer fuente y tamaño
            pdf.set_font("Arial", style='B', size=14)
        
        # Título del PDF
            pdf.cell(200, 10, txt="Libreta de Contactos", ln=True, align="C")
        
        # Salto de línea
            pdf.ln(10)
        
        # Establecer fuente para los datos de la tabla
            pdf.set_font("Arial", size=12)
        
        # Encabezados de la tabla
            pdf.set_fill_color(200, 220, 255)
            pdf.cell(60, 10, "Nombre", border=1, align="C", fill=True)
            pdf.cell(50, 10, "Teléfono", border=1, align="C", fill=True)
            pdf.cell(80, 10, "Correo", border=1, align="C", fill=True)
            pdf.ln()

        # Escribir los contactos en el PDF
            with open("libreta_contactos.csv", "r", newline='', encoding="utf-8") as f:
                for nombre, telefono, correo in csv.reader(f):
                    pdf.cell(60, 10, nombre, border=1, align="C")
                    pdf.cell(50, 10, telefono, border=1, align="C")
                    pdf.cell(80, 10, correo, border=1, align="C")
                    pdf.ln()

        # Guardar el archivo PDF
            pdf.output(archivo_pdf)
            messagebox.showinfo("PDF descargado", f"El archivo PDF ha sido descargado en: {archivo_pdf}")
        except Exception as e:
         messagebox.showerror("Error", f"Error al generar el PDF: {e}")



    def enviar_correo_contacto(self):
        try:
            correo_destino = self.correo_entry.get()
            if not correo_destino:
                messagebox.showwarning("Correo vacío", "Por favor, ingrese un correo de destino.")
                return
            servidor = smtplib.SMTP("smtp.gmail.com", 587)
            servidor.starttls()
            servidor.login("riveramelvin628@gmail.com", "xtqm uagl sjpv btew")
            
            mensaje = MIMEMultipart()
            mensaje["From"] = "riveramelvin628@gmail.com"
            mensaje["To"] = correo_destino
            mensaje["Subject"] = "Contacto desde la Libreta"
            
            cuerpo = f"Hola, este es un mensaje de prueba desde la libreta de contactos.\n\nNombre: {self.nombre_entry.get()}\nTeléfono: {self.telefono_entry.get()}"
            mensaje.attach(MIMEText(cuerpo, "plain"))
            
            servidor.sendmail("riveramelvin628@gmail.com", correo_destino, mensaje.as_string())
            servidor.quit()
            messagebox.showinfo("Correo enviado", f"Correo enviado a {correo_destino}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al enviar el correo: {e}")

if __name__ == "__main__":
    root = ttkb.Window(themename="superhero")
    app = LibretaContactosApp(root)
    app.mostrar_contactos()
    root.mainloop()
